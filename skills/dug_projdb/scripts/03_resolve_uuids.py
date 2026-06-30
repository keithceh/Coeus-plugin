"""
DUG project DB → inventory xlsx (step 3 of 3).

Resolves UUIDs inside Volume_Processes / Process_Catalog parameter strings to
human-readable product names. Edits xlsx in place.

DUG UUID encoding: Product.productUUIDMSB / productUUIDLSB are signed int64
columns holding the top/bottom 64 bits of the 128-bit UUID. Two's-complement
conversion required for SQLite signed→unsigned int64.

Adds a UUID_Index sheet — every UUID found in any parameter string with its
resolved product info (or "(not found)" for the rare horizon-property UUIDs
that aren't first-class Products).

Usage:
    python 03_resolve_uuids.py --db path/to/project.dugprj --out path/to/out.xlsx
"""
import argparse, sqlite3, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

UUID_RE = re.compile(r"\b[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}\b")

KIND = {
    1:"flt", 2:"well", 3:"pick", 4:"vsp", 5:"log", 6:"curve", 7:"pickset",
    8:"mrk", 9:"proj", 10:"2dsv", 11:"2dln", 12:"hor", 13:"vol", 14:"fld",
    17:"proc", 18:"poly", 19:"xsec", 23:"wav", 28:"wavf", 29:"?29", 30:"?30",
    31:"prj", 37:"shp", 40:"crvm", 41:"litho", 43:"seq", 46:"loop",
    47:"txt", 49:"scn", 51:"drad", 52:"ovly", 53:"wdelta", 54:"reg",
}


def u64_to_s64(x): return x - 0x10000000000000000 if x >= 0x8000000000000000 else x
def s64_to_u64(x): return x + 0x10000000000000000 if x < 0 else x


def msb_lsb_to_uuid(msb, lsb):
    msb, lsb = s64_to_u64(msb), s64_to_u64(lsb)
    h = f"{msb:016x}{lsb:016x}"
    return f"{h[0:8]}-{h[8:12]}-{h[12:16]}-{h[16:20]}-{h[20:32]}"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    con = sqlite3.connect(f"file:{args.db}?mode=ro", uri=True); cur = con.cursor()
    cur.execute("""SELECT p.productId, p.productUUIDMSB, p.productUUIDLSB, p.productType,
                          (SELECT value FROM AttributeRevision
                           WHERE productId=p.productId AND attributeType=1 AND value IS NOT NULL AND value!='' LIMIT 1) AS name
                   FROM Product p""")
    uuid_map = {}
    for pid, msb, lsb, ptype, name in cur.fetchall():
        if msb is None or lsb is None: continue
        uuid_map[msb_lsb_to_uuid(msb, lsb)] = (pid, ptype, name or "(unnamed)")
    print(f"built UUID map: {len(uuid_map)} entries")

    wb = load_workbook(args.out)

    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    HDR_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    BORDER = Border(left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
                    top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"))
    ALIGN_TOP = Alignment(vertical="top", wrap_text=True)

    def resolve(u):
        info = uuid_map.get(u)
        if not info: return None
        pid, ptype, name = info
        kind = KIND.get(ptype, f"t{ptype}")
        return f"«{name}» ({kind}/pid{pid})"

    found = {}; sub_count = 0
    def substitute(s):
        nonlocal sub_count
        if not s: return s
        def repl(m):
            nonlocal sub_count
            u = m.group(0); r = resolve(u)
            if u not in found: found[u] = r
            if r is None: return u
            sub_count += 1; return r
        return UUID_RE.sub(repl, s)

    for sheet_name in ("Volume_Processes", "Process_Catalog"):
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        col = 7
        for row in range(3, ws.max_row+1):
            c = ws.cell(row=row, column=col)
            if c.value: c.value = substitute(c.value)

    if "UUID_Index" in wb.sheetnames:
        del wb["UUID_Index"]
    ws_u = wb.create_sheet("UUID_Index")
    unresolved = sum(1 for v in found.values() if v is None)
    note = (f"Every UUID encountered in parameter strings. "
            f"{len(found) - unresolved} of {len(found)} resolved to a Product.")
    c = ws_u.cell(row=1, column=1, value=note)
    c.font = Font(italic=True, color="7F6000"); c.fill = WARN_FILL
    ws_u.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    headers = ["UUID", "resolved name", "kind", "productId", "productType"]
    for ci, h in enumerate(headers, 1):
        cc = ws_u.cell(row=2, column=ci, value=h)
        cc.font = HDR_FONT; cc.fill = HDR_FILL
        cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cc.border = BORDER
    ws_u.row_dimensions[2].height = 30
    ri = 3
    for u in sorted(found.keys()):
        info = uuid_map.get(u)
        if info:
            pid, ptype, name = info
            vals = [u, name, KIND.get(ptype, f"t{ptype}"), pid, ptype]
        else:
            vals = [u, "(not found in Product table)", "", "", ""]
        for ci, v in enumerate(vals, 1):
            cc = ws_u.cell(row=ri, column=ci, value=v); cc.alignment = ALIGN_TOP; cc.border = BORDER
        ri += 1
    for ci, w in enumerate([40, 60, 8, 12, 14], 1):
        ws_u.column_dimensions[get_column_letter(ci)].width = w
    ws_u.freeze_panes = "A3"

    ws_r = wb["README"]
    row = ws_r.max_row + 2
    ws_r.cell(row=row, column=1, value="UPDATE — UUID resolution").font = Font(bold=True, color="1F4E78"); row += 1
    ws_r.cell(row=row, column=1, value="  UUIDs encountered in parameter strings").alignment = ALIGN_TOP
    ws_r.cell(row=row, column=2, value=len(found)).alignment = ALIGN_TOP; row += 1
    ws_r.cell(row=row, column=1, value="  UUIDs resolved").alignment = ALIGN_TOP
    ws_r.cell(row=row, column=2, value=len(found) - unresolved).alignment = ALIGN_TOP; row += 1
    ws_r.cell(row=row, column=1, value="  UUIDs unresolved (kept verbatim)").alignment = ALIGN_TOP
    ws_r.cell(row=row, column=2, value=unresolved).alignment = ALIGN_TOP; row += 1
    ws_r.cell(row=row, column=1, value="  Total substitutions applied").alignment = ALIGN_TOP
    ws_r.cell(row=row, column=2, value=sub_count).alignment = ALIGN_TOP; row += 1
    ws_r.cell(row=row, column=1, value="  Substitution format").alignment = ALIGN_TOP
    ws_r.cell(row=row, column=2, value="«name» (kind/pidN)").alignment = ALIGN_TOP

    desired = ["README","Volumes","Volume_Processes","Process_Catalog","Horizons","Polygons","ProductType_Map","UUID_Index"]
    wb._sheets = [wb[n] for n in desired if n in wb.sheetnames]
    wb.save(args.out)
    print(f"OK step-3 → {args.out}")
    print(f"  {len(found)} distinct UUIDs found, {len(found)-unresolved} resolved, {unresolved} unresolved")
    print(f"  {sub_count} substitutions applied")


if __name__ == "__main__":
    main()
