"""
DUG inventory xlsx → interactive volume-lineage HTML explorer (step 5 of 5).

Builds the parent/child volume graph from «name» (vol/pidN) references inside
Volume_Processes parameter strings (self-loops skipped), splits it into
weakly-connected components, lays out each non-trivial component (size >= 2)
top-down, and writes ONE self-contained HTML file: no external scripts,
styles, fonts, or images. Features: component tabs, hover-to-enlarge,
click-to-pin with parent (blue) / child (orange) highlighting + dimming,
sidebar with clickable parent/child lists, search, whole-diagram pan+zoom.

Hardened against "blank page" failure modes: <noscript> notice and a
window.onerror banner — a blocked or failed script shows a visible message
instead of nothing.

Usage:
    python 05_lineage_explorer.py --xlsx path/to/inventory.xlsx --out path/to/explorer.html [--title "My Project"]
    python 05_lineage_explorer.py --selftest
"""
import argparse, html, json, re, sys
from collections import defaultdict

PID_REF = re.compile(r"«([^»]+)» \(vol/pid(\d+)\)")

# ------------------------------------------------------------------ graph

def load_rows(xlsx):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Volume_Processes"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    wb.close()
    return rows


def build_graph(rows):
    """rows: (vpid, vname, dpid, dtype, ptype, dname, params).
    Returns (names {pid:name}, edges set[(parent_pid, child_pid)])."""
    names, edges = {}, set()
    for vpid, vname, dpid, dtype, ptype, dname, params in rows:
        if vpid is not None:
            names.setdefault(int(vpid), vname or f"(pid {vpid})")
        if dpid in (None, "—"):
            continue
        for m in PID_REF.finditer(params or ""):
            ppid = int(m.group(2))
            if ppid == int(vpid):
                continue  # self-loop
            names.setdefault(ppid, m.group(1))
            edges.add((ppid, int(vpid)))
    return names, edges


def components(names, edges):
    """Weakly-connected components, largest first. Only size >= 2 returned."""
    parent = {p: p for p in names}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    for a, b in edges:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb
    groups = defaultdict(list)
    for p in names:
        groups[find(p)].append(p)
    comps = [sorted(g) for g in groups.values() if len(g) >= 2]
    comps.sort(key=len, reverse=True)
    return comps


# ------------------------------------------------------------------ layout

def layer_nodes(nodes, edges):
    """Longest-path layering, cycle-tolerant (Kahn; leftovers appended)."""
    nodes = set(nodes)
    out = defaultdict(set)
    indeg = {n: 0 for n in nodes}
    for a, b in edges:
        if a in nodes and b in nodes and b not in out[a]:
            out[a].add(b)
            indeg[b] += 1
    layer = {n: 0 for n in nodes}
    queue = [n for n in nodes if indeg[n] == 0]
    seen = set(queue)
    while queue:
        n = queue.pop(0)
        for c in out[n]:
            layer[c] = max(layer[c], layer[n] + 1)
            indeg[c] -= 1
            if indeg[c] == 0:
                queue.append(c); seen.add(c)
    # ponytail: cycle members (indeg never reaches 0) keep layer 0 defaults
    # bumped below their parents where possible; proper cycle-breaking (FAS)
    # if a project ever shows tangled cycles.
    for n in nodes - seen:
        preds = [p for p, c in edges if c == n and p in seen]
        if preds:
            layer[n] = max(layer[p] for p in preds) + 1
    return layer


def layout_component(nodes, all_edges, names):
    """Returns (svg_string, width, height) for one component."""
    edges = [(a, b) for a, b in all_edges if a in nodes and b in nodes]
    layer = layer_nodes(nodes, edges)
    by_layer = defaultdict(list)
    for n in nodes:
        by_layer[layer[n]].append(n)

    def label(pid):
        n = str(names.get(pid, pid))
        return n if len(n) <= 35 else n[:35] + "…"

    CHAR_W, PAD_X, NODE_H, GAP_X, GAP_Y, MARGIN = 6.1, 14, 36, 18, 72, 12
    width_of = {n: max(70, len(label(n)) * CHAR_W + 2 * PAD_X) for n in nodes}

    # order within layer by mean parent x-order (one barycenter pass)
    order = {}
    for ly in sorted(by_layer):
        row = by_layer[ly]
        if ly == 0:
            row.sort(key=lambda n: str(names.get(n, "")).lower())
        else:
            prev = order
            row.sort(key=lambda n: (
                sum(prev.get(p, 0) for p, c in edges if c == n) /
                max(1, sum(1 for p, c in edges if c == n))))
        for i, n in enumerate(row):
            order[n] = i
        by_layer[ly] = row

    row_widths = {ly: sum(width_of[n] for n in row) + GAP_X * (len(row) - 1)
                  for ly, row in by_layer.items()}
    total_w = max(row_widths.values()) + 2 * MARGIN
    total_h = (max(by_layer) + 1) * GAP_Y + NODE_H + 2 * MARGIN

    pos = {}
    for ly, row in by_layer.items():
        x = (total_w - row_widths[ly]) / 2
        y = MARGIN + ly * GAP_Y
        for n in row:
            pos[n] = (x, y)
            x += width_of[n] + GAP_X

    parents_in = {n: [a for a, b in edges if b == n] for n in nodes}
    children_in = {n: [b for a, b in edges if a == n] for n in nodes}

    def fill(n):
        if not parents_in[n]:
            return "#B8CCE4"   # root
        if not children_in[n]:
            return "#FBE5D6"   # leaf
        return "#DDEBF7"       # mid-chain

    parts = []
    for a, b in edges:
        ax, ay = pos[a]; bx, by = pos[b]
        x1, y1 = ax + width_of[a] / 2, ay + NODE_H
        x2, y2 = bx + width_of[b] / 2, by
        dx, dy = x2 - x1, y2 - y1
        ln = max(1e-6, (dx * dx + dy * dy) ** 0.5)
        ux, uy = dx / ln, dy / ln
        tipx, tipy = x2, y2
        bxp, byp = tipx - ux * 9, tipy - uy * 9
        px, py = -uy * 3.5, ux * 3.5
        parts.append(
            f'<g id="e{a}_{b}" class="edge">'
            f'<path fill="none" stroke="#1f4e78" stroke-width="1.1" '
            f'd="M{x1:.1f},{y1:.1f}L{bxp:.1f},{byp:.1f}"/>'
            f'<polygon fill="#1f4e78" stroke="#1f4e78" points="'
            f'{tipx:.1f},{tipy:.1f} {bxp + px:.1f},{byp + py:.1f} {bxp - px:.1f},{byp - py:.1f}"/></g>')
    for n in nodes:
        x, y = pos[n]
        w = width_of[n]
        parts.append(
            f'<g id="n{n}" class="node">'
            f'<title>{html.escape(str(names.get(n, n)))}</title>'
            f'<rect x="{x:.1f}" y="{y:.1f}" width="{w:.1f}" height="{NODE_H}" rx="10" '
            f'fill="{fill(n)}" stroke="#1f4e78" stroke-width="1.2"/>'
            f'<text x="{x + w / 2:.1f}" y="{y + NODE_H / 2 + 4:.1f}" text-anchor="middle" '
            f'font-family="Calibri,Arial,sans-serif" font-size="11">{html.escape(label(n))}</text></g>')

    svg = (f'<svg width="{total_w:.0f}" height="{total_h:.0f}" '
           f'viewBox="0 0 {total_w:.0f} {total_h:.0f}" '
           f'xmlns="http://www.w3.org/2000/svg">'
           f'<g id="graph0">{"".join(parts)}</g></svg>')
    return svg, total_w, total_h


# ------------------------------------------------------------------ html

TEMPLATE = """<!doctype html>
<html lang="en"><head>
<meta charset="utf-8">
<title>Volume Lineage Explorer — __TITLE__</title>
<style>
:root{--bg:#0f1116;--fg:#e6e7ea;--mut:#9aa0a6;--acc:#f4b400;--card:#161922;--bd:#262a36;--blue:#1F4E78;}
*{box-sizing:border-box}
html,body{margin:0;padding:0;background:var(--bg);color:var(--fg);font:14px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Arial,sans-serif;}
header{position:sticky;top:0;background:rgba(15,17,22,.95);backdrop-filter:blur(8px);border-bottom:1px solid var(--bd);padding:14px 24px;z-index:100;}
header h1{margin:0 0 4px;font-size:18px;letter-spacing:.3px;}
header .sub{color:var(--mut);font-size:13px;}
.errbar{display:none;background:#7f1d1d;color:#fff;padding:10px 24px;font-size:13px;}
noscript .errbar{display:block;}
main{display:grid;grid-template-columns:1fr 320px;gap:0;min-height:calc(100vh - 64px);}
.viewport{padding:16px;overflow:auto;}
.tabs{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:14px;padding:0;}
.tab{padding:6px 12px;background:var(--card);border:1px solid var(--bd);border-radius:6px;color:var(--mut);cursor:pointer;font-size:13px;}
.tab.active{color:var(--acc);border-color:var(--acc);background:#1e2431;}
.tab:hover:not(.active){color:var(--fg);border-color:var(--fg);}
.legend{margin-bottom:14px;padding:10px 14px;background:var(--card);border:1px solid var(--bd);border-radius:6px;font-size:13px;color:var(--mut);}
.legend .swatch{display:inline-block;width:14px;height:14px;border-radius:3px;vertical-align:middle;margin:0 6px 0 12px;border:1px solid var(--bd);}
.hint{font-size:12px;color:var(--mut);margin-top:6px;}
.svg-frame{background:#ffffff;border:1px solid var(--bd);border-radius:8px;overflow:hidden;position:relative;min-height:400px;}
.svg-frame svg{display:block;width:100%;height:auto;cursor:grab;max-height:75vh;}
.svg-frame svg:active{cursor:grabbing;}
.svg-frame g.node{cursor:pointer;transition:transform .18s ease-out,filter .18s;transform-box:fill-box;transform-origin:center;}
.svg-frame g.node:hover{transform:scale(1.35);filter:drop-shadow(0 4px 12px rgba(31,78,120,.35));}
.svg-frame g.node:hover polygon,.svg-frame g.node:hover rect,.svg-frame g.node:hover path{stroke:#f4b400 !important;stroke-width:2 !important;}
.svg-frame g.node.pinned polygon,.svg-frame g.node.pinned rect{stroke:#c55a11 !important;stroke-width:3 !important;}
.svg-frame g.node.parent-of polygon,.svg-frame g.node.parent-of rect{stroke:#2E75B6 !important;stroke-width:2.5 !important;}
.svg-frame g.node.child-of polygon,.svg-frame g.node.child-of rect{stroke:#C55A11 !important;stroke-width:2.5 !important;}
.svg-frame g.node.dim{opacity:.25;}
.svg-frame g.edge.dim{opacity:.1;}
.zoom-ctrl{position:absolute;top:10px;right:10px;display:flex;flex-direction:column;gap:4px;z-index:20;}
.zoom-ctrl button{width:32px;height:32px;border:1px solid var(--bd);background:rgba(255,255,255,.95);color:var(--blue);border-radius:6px;cursor:pointer;font-size:16px;font-weight:bold;}
.zoom-ctrl button:hover{background:#eee;}
aside{background:var(--card);border-left:1px solid var(--bd);padding:16px 18px;overflow-y:auto;max-height:calc(100vh - 64px);position:sticky;top:64px;}
aside h2{margin:0 0 6px;font-size:15px;color:var(--acc);}
aside .name{font-weight:600;color:var(--fg);word-break:break-word;font-size:15px;margin-bottom:12px;}
aside .empty{color:var(--mut);font-style:italic;padding:24px 0;text-align:center;font-size:13px;}
aside h3{margin:14px 0 6px;font-size:13px;color:var(--mut);text-transform:uppercase;letter-spacing:.5px;}
aside .parents h3{color:#7cb1ff;} aside .children h3{color:#f6a561;}
aside ul{list-style:none;margin:0;padding:0;}
aside li{padding:6px 8px;border-radius:4px;font-size:13px;margin-bottom:2px;cursor:pointer;word-break:break-word;background:#1c2130;}
aside li:hover{background:#252a3a;color:var(--acc);}
aside .stats{color:var(--mut);font-size:12px;margin-top:20px;padding-top:12px;border-top:1px solid var(--bd);}
.search{width:100%;padding:8px 10px;background:#1c2130;border:1px solid var(--bd);border-radius:4px;color:var(--fg);font-size:13px;margin-bottom:8px;}
@media (max-width: 900px){ main{grid-template-columns:1fr;} aside{border-left:0;border-top:1px solid var(--bd);position:static;} }
</style></head><body>
<header>
  <h1>Volume Lineage Explorer — __TITLE__</h1>
  <div class="sub">__STATS__. Hover a node to enlarge. Click to pin.</div>
</header>
<noscript><div class="errbar" style="display:block">This explorer needs JavaScript. Your viewer has scripts disabled — open the file in a normal browser (Edge, Chrome, Firefox).</div></noscript>
<div class="errbar" id="errbar"></div>
<main>
  <div class="viewport">
    <div class="legend">
      <span class="swatch" style="background:#B8CCE4"></span>root (no parents)
      <span class="swatch" style="background:#DDEBF7"></span>mid-chain
      <span class="swatch" style="background:#FBE5D6"></span>leaf (no children)
      <div class="hint">Hover a node → it enlarges. Click → parents highlight blue, children orange, unrelated nodes dim. Click empty space to clear. Scroll or +/− to zoom the whole diagram; drag to pan.</div>
    </div>
    <div class="tabs" id="tabs"></div>
    <div class="svg-frame" id="frame">
      <div class="zoom-ctrl">
        <button id="zoomIn" title="Zoom in">+</button>
        <button id="zoomOut" title="Zoom out">-</button>
        <button id="zoomReset" title="Reset" style="font-size:12px">R</button>
      </div>
      <div id="svgHost"></div>
    </div>
  </div>
  <aside>
    <input type="text" class="search" id="search" placeholder="Search volume name...">
    <h2>Selected Volume</h2>
    <div id="detail"><div class="empty">Click a node to see its parents and children.</div></div>
  </aside>
</main>
<script>
window.onerror = function (msg) {
  var e = document.getElementById('errbar');
  if (e) { e.style.display = 'block'; e.textContent = 'Explorer script error: ' + msg; }
};
const COMPONENTS = __COMPONENTS__;
const NODES = __NODES__;

const SVG_NS = 'http://www.w3.org/2000/svg';
const tabsEl = document.getElementById('tabs');
const host = document.getElementById('svgHost');
const detail = document.getElementById('detail');
let currentIdx = 0, pinnedPid = null, zoom = 1, panX = 0, panY = 0;
let userTx = null;

COMPONENTS.forEach((c, i) => {
  const t = document.createElement('div');
  t.className = 'tab' + (i === 0 ? ' active' : '');
  t.textContent = 'Component ' + c.idx + ' - ' + c.size;
  t.addEventListener('click', () => switchComp(i));
  tabsEl.appendChild(t);
});

function switchComp(i) {
  currentIdx = i;
  document.querySelectorAll('.tab').forEach((t, j) => t.classList.toggle('active', j === i));
  host.innerHTML = COMPONENTS[i].svg;
  const svg = host.querySelector('svg');
  if (!svg) return;
  const graph0 = svg.querySelector('g#graph0') || svg.querySelector('g');
  if (!graph0) return;
  const wrapper = document.createElementNS(SVG_NS, 'g');
  wrapper.setAttribute('id', 'userTx');
  while (graph0.firstChild) { wrapper.appendChild(graph0.firstChild); }
  graph0.appendChild(wrapper);
  userTx = wrapper;
  zoom = 1; panX = 0; panY = 0; applyTransform();
  bindNodes(); clearPin();
}

function bindNodes() {
  const svg = host.querySelector('svg'); if (!svg) return;
  svg.querySelectorAll('g.node').forEach(g => {
    const pid = g.id.replace(/^n/, '');
    g.addEventListener('click', ev => { ev.stopPropagation(); pinNode(pid); });
  });
  svg.addEventListener('click', () => clearPin());
  enablePan(svg); enableWheelZoom(svg);
}

function pinNode(pid) {
  pinnedPid = pid;
  const meta = NODES[pid]; if (!meta) { clearPin(); return; }
  const parentPids = new Set(), childPids = new Set();
  meta.parents.forEach(n => { for (const [k, v] of Object.entries(NODES)) if (v.name === n) parentPids.add(k); });
  meta.children.forEach(n => { for (const [k, v] of Object.entries(NODES)) if (v.name === n) childPids.add(k); });
  document.querySelectorAll('g.node').forEach(g => {
    const p = g.id.replace(/^n/,'');
    g.classList.remove('pinned','parent-of','child-of','dim');
    if (p === pid) g.classList.add('pinned');
    else if (parentPids.has(p)) g.classList.add('parent-of');
    else if (childPids.has(p)) g.classList.add('child-of');
    else g.classList.add('dim');
  });
  document.querySelectorAll('g.edge').forEach(g => {
    const m = g.id.match(/^e(\\d+)_(\\d+)$/); if (!m) return;
    const involved = (m[1] === pid || m[2] === pid);
    g.classList.toggle('dim', !involved);
  });
  renderDetail(meta);
}
function clearPin() {
  pinnedPid = null;
  document.querySelectorAll('g.node').forEach(g => g.classList.remove('pinned','parent-of','child-of','dim'));
  document.querySelectorAll('g.edge').forEach(g => g.classList.remove('dim'));
  detail.innerHTML = '<div class="empty">Click a node to see its parents and children.</div>';
}
function renderDetail(meta) {
  const item = n => '<li data-name="' + esc(n) + '">' + esc(n) + '</li>';
  detail.innerHTML =
    '<div class="name">' + esc(meta.name) + '</div>' +
    '<div class="parents"><h3>Parents (' + meta.parents.length + ')</h3>' +
    (meta.parents.length ? '<ul>' + meta.parents.map(item).join('') + '</ul>' : '<div class="empty" style="padding:8px 0">- none -</div>') + '</div>' +
    '<div class="children"><h3>Children (' + meta.children.length + ')</h3>' +
    (meta.children.length ? '<ul>' + meta.children.map(item).join('') + '</ul>' : '<div class="empty" style="padding:8px 0">- none -</div>') + '</div>' +
    '<div class="stats">Click a name to jump to that volume.</div>';
  detail.querySelectorAll('li').forEach(li => {
    li.addEventListener('click', () => {
      const name = li.dataset.name;
      const pid = Object.keys(NODES).find(k => NODES[k].name === name);
      if (!pid) return;
      const g = document.getElementById('n' + pid);
      if (!g) {
        const foundIdx = COMPONENTS.findIndex(c => c.svg.includes('id="n' + pid + '"'));
        if (foundIdx >= 0) { switchComp(foundIdx); requestAnimationFrame(() => pinNode(pid)); }
      } else { pinNode(pid); g.scrollIntoView({behavior:'smooth', block:'center'}); }
    });
  });
}
function esc(s) { return String(s).replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c])); }

function applyTransform() {
  if (!userTx) return;
  userTx.setAttribute('transform', 'translate(' + panX + ',' + panY + ') scale(' + zoom + ')');
}
function enablePan(svg) {
  let dragging = false, sX = 0, sY = 0, sPX = 0, sPY = 0;
  svg.addEventListener('mousedown', e => {
    if (e.target.closest('g.node')) return;
    dragging = true; sX = e.clientX; sY = e.clientY; sPX = panX; sPY = panY;
  });
  window.addEventListener('mousemove', e => {
    if (!dragging) return;
    panX = sPX + (e.clientX - sX); panY = sPY + (e.clientY - sY); applyTransform();
  });
  window.addEventListener('mouseup', () => dragging = false);
}
function enableWheelZoom(svg) {
  svg.addEventListener('wheel', e => {
    e.preventDefault();
    zoom = Math.max(0.3, Math.min(5, zoom * (e.deltaY < 0 ? 1.1 : 1/1.1)));
    applyTransform();
  }, { passive: false });
}
document.getElementById('zoomIn').addEventListener('click', () => { zoom = Math.min(5, zoom*1.2); applyTransform(); });
document.getElementById('zoomOut').addEventListener('click', () => { zoom = Math.max(0.3, zoom/1.2); applyTransform(); });
document.getElementById('zoomReset').addEventListener('click', () => { zoom = 1; panX = 0; panY = 0; applyTransform(); });

document.getElementById('search').addEventListener('input', e => {
  const q = e.target.value.trim().toLowerCase();
  if (q.length < 2) return;
  const match = Object.entries(NODES).find(([pid, m]) => m.name.toLowerCase().includes(q));
  if (!match) return;
  const [pid] = match;
  const foundIdx = COMPONENTS.findIndex(c => c.svg.includes('id="n' + pid + '"'));
  if (foundIdx >= 0) {
    if (foundIdx !== currentIdx) switchComp(foundIdx);
    requestAnimationFrame(() => {
      pinNode(pid);
      const g = document.getElementById('n' + pid);
      if (g) g.scrollIntoView({behavior:'smooth', block:'center'});
    });
  }
});

if (!COMPONENTS.length) {
  host.innerHTML = '<div style="padding:40px;color:#333">No non-trivial lineage components found in this project.</div>';
} else {
  switchComp(0);
}
</script></body></html>
"""


def build_html(names, edges, title):
    comps = components(names, edges)
    comp_objs = []
    for i, comp in enumerate(comps, 1):
        svg, _, _ = layout_component(set(comp), edges, names)
        comp_objs.append({"idx": i, "size": len(comp), "svg": svg})

    parents = defaultdict(list)
    children = defaultdict(list)
    for a, b in sorted(edges):
        parents[b].append(names[a])
        children[a].append(names[b])
    nodes = {str(p): {"name": str(names[p]),
                      "parents": sorted(parents[p], key=str.lower),
                      "children": sorted(children[p], key=str.lower)}
             for p in names}

    n_with_parents = sum(1 for p in names if parents[p])
    stats = (f"{len(names)} volumes · {len(edges)} parent-child edges · "
             f"{len(comp_objs)} non-trivial components")
    page = (TEMPLATE
            .replace("__TITLE__", html.escape(title))
            .replace("__STATS__", stats)
            .replace("__COMPONENTS__", json.dumps(comp_objs))
            .replace("__NODES__", json.dumps(nodes, ensure_ascii=False)))
    assert "src=" not in page.split("<script>")[0], "template must stay self-contained"
    return page, len(comp_objs), n_with_parents


def selftest():
    names = {1: "A", 2: "B", 3: "C", 4: "loner", 5: "X", 6: "Y"}
    edges = {(1, 2), (2, 3), (5, 6)}
    comps = components(names, edges)
    assert [len(c) for c in comps] == [3, 2], comps  # loner excluded
    layer = layer_nodes({1, 2, 3}, edges)
    assert (layer[1], layer[2], layer[3]) == (0, 1, 2)
    # cycle tolerance
    lc = layer_nodes({1, 2}, {(1, 2), (2, 1)})
    assert set(lc) == {1, 2}
    page, ncomp, _ = build_html(names, edges, "T & T")
    assert ncomp == 2 and 'id=\\"n1\\"' in page and "T &amp; T" in page
    assert page.count("<script>") == 1 and "http" not in page.split("</style>")[1].split("<script>")[0]
    print("selftest OK")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", help="step-3 inventory workbook")
    ap.add_argument("--out", help="output html path")
    ap.add_argument("--title", default=None, help="project title shown in the header")
    ap.add_argument("--selftest", action="store_true")
    a = ap.parse_args()
    if a.selftest:
        selftest(); sys.exit(0)
    if not (a.xlsx and a.out):
        ap.error("--xlsx and --out are required (no defaults — user must supply paths)")
    title = a.title or __import__("pathlib").Path(a.xlsx).stem
    names, edges = build_graph(load_rows(a.xlsx))
    page, ncomp, nwp = build_html(names, edges, title)
    with open(a.out, "w", encoding="utf-8") as f:
        f.write(page)
    print(f"OK -> {a.out}  ({len(names)} volumes, {len(edges)} edges, "
          f"{ncomp} non-trivial components, {nwp} volumes with >=1 parent)")
