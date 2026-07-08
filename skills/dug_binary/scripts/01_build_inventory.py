"""
DUG project DB -> inventory xlsx (step 1 of 3).

Reads a DUG Insight project.dugprj (SQLite, schema v16.x) and produces an
Excel inventory with sheets: README, Volumes, Volume_Processes (lineage only,
no parameters yet — step 2 enriches), Horizons, Polygons, ProductType_Map.

The DUG schema is undocumented. productType->category mapping is INFERRED
from filenames (120hors/, 100sei/, 191culture/) and sample attribute values.
Verify with the project owner before high-stakes use.

Usage:
    python 01_build_inventory.py --db path/to/project.dugprj --out path/to/out.xlsx
"""
import argparse, sqlite3, datetime as dt, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# productType -> (category, label) — inferred mapping
PT_MAP = {
    1:  ("Fault sticks",          "Fault stick set / unassigned sticks"),
    2:  ("Well",                  "Well object"),
    3:  ("Well pre-pick / pick set", "Pre-pick or pick set imported from txt"),
    4:  ("VSP velocity",          "VSP velocity / checkshot data"),
    5:  ("Well log (wireline)",   "Wireline / LWD curve set"),
    6:  ("Well log curve",        "Individual log curve (DEPTH, CALI, DEN, ...)"),
    7:  ("Pick set",              "Named pick set / formation top collection"),
    8:  ("Horizon marker / pick name", "Marker / formation-top label"),
    9:  ("Survey / processing project", "Processing project container"),
    10: ("2D Survey",             "2D survey container"),
    11: ("2D seismic line",       "Individual 2D line"),
    12: ("HORIZON (gridded surface)", "Interpreted horizon (TWT or depth grid)"),
    13: ("VOLUME (seismic)",      "3D seismic volume or 2D survey lines container"),
    14: ("Folder / group",        "Organising folder"),
    17: ("Workflow / process",    "Process step (filter, attribute, math, conditioning)"),
    18: ("POLYGON / contour / outline", "Map polygon, depth contour, AOI, outline"),
    19: ("Cross-section line",    "Line between points"),
    23: ("Wavelet",               "Wavelet (Ricker / extracted / synthetic)"),
    28: ("Wavelet file",          "On-disk wavelet (.su)"),
    29: ("Property / axis descriptor", "Property slot on horizons/polygons/2D-lines/cross-sections (attr 201 = slot index; inferred 2026-Jul-03)"),
    30: ("Individual fault stick", "One per stick; referenced by fault-stick sets (type 1) via attr 400 (inferred 2026-Jul-03)"),
    31: ("Project",               "Project root entry"),
    37: ("Culture shapefile",     "Imported shapefile (blocks, basin outlines)"),
    40: ("Curve maths",           "Synthetic / computed log curve"),
    41: ("Litho model",           "Lithology model"),
    43: ("Sequence",              "Stratigraphic sequence"),
    46: ("Loop",                  "Workflow loop / iteration"),
    47: ("Text overlay",          "Imported text overlay (.txt) on data"),
    49: ("Scenario",              "Scenario container"),
    51: ("Drilling radius",       "Circular drilling-radius polygon"),
    52: ("Velocity / inversion overlay", "Vel / inversion overlay (linked to a source volume)"),
    53: ("Wells delta table",     "Inter-horizon thickness table"),
    54: ("Regression model",      "Regression model"),
}

# AttributeRevision.attributeType (empirical)
AT_NAME, AT_DESC, AT_OWNER, AT_HASH, AT_PATH, AT_CRS = 1, 3, 4, 8, 7, 10
AT_UNIT, AT_MEASURE = 102, 101


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", required=True, help="path to project.dugprj")
    ap.add_argument("--out", required=True, help="output xlsx path")
    args = ap.parse_args()

    con = sqlite3.connect(f"file:{args.db}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    def attrs_for(pids, attribute_types):
        out = {}
        if not pids: return out
        ph_p = ",".join("?"*len(pids)); ph_a = ",".join("?"*len(attribute_types))
        cur.execute(f"""SELECT productId, attributeType, value FROM AttributeRevision
                        WHERE productId IN ({ph_p}) AND attributeType IN ({ph_a})""",
                    list(pids)+list(attribute_types))
        for pid, at, v in cur.fetchall():
            d = out.setdefault(pid, {})
            if at not in d or (d[at] is None and v is not None):
                d[at] = v
        return out

    def epoch_ms(x):
        if x is None: return None
        try: return dt.datetime.utcfromtimestamp(int(x)/1000).strftime("%Y-%m-%d %H:%M:%S")
        except Exception: return None

    def fetch_products(pt):
        cur.execute("SELECT productId, productType, owner FROM Product WHERE productType=?", (pt,))
        return cur.fetchall()

    horizons_products = fetch_products(12)
    markers_products  = fetch_products(8)
    polygons_products = fetch_products(18)
    culture_products  = fetch_products(37)
    volumes_products  = fetch_products(13)
    workflow_products = fetch_products(17)
    velovly_products  = fetch_products(52)

    vol_pids = [r["productId"] for r in volumes_products]
    cur.execute(f"""SELECT pr.productId, COUNT(DISTINCT pr.revisionId) AS n_rev,
                           MIN(r.revisionTime), MAX(r.revisionTime), COUNT(DISTINCT r.user)
                    FROM ProductRevision pr JOIN Revision r ON r.revisionId=pr.revisionId
                    WHERE pr.productId IN ({",".join("?"*len(vol_pids))}) GROUP BY pr.productId""",
                vol_pids)
    vol_rev = {row[0]: {"n_rev": row[1], "first_t": row[2], "last_t": row[3], "n_users": row[4]}
               for row in cur.fetchall()}

    cur.execute("""SELECT vh.volumeUUIDMSB, vh.volumeUUIDLSB, COUNT(*),
                          MAX(vh.trace_count), MAX(vh.trace_ns)
                   FROM volume_headers vh GROUP BY vh.volumeUUIDMSB, vh.volumeUUIDLSB""")
    vh_by_uuid = {(r[0], r[1]): r for r in cur.fetchall()}

    cur.execute("SELECT productId, productUUIDMSB, productUUIDLSB FROM Product WHERE productType=13")
    vol_uuid_by_pid = {r[0]: (r[1], r[2]) for r in cur.fetchall()}

    cur.execute("""SELECT DISTINCT ar.relatedProductId, ar.productId
                   FROM AttributeRevision ar JOIN Product src_p ON src_p.productId=ar.relatedProductId
                   WHERE src_p.productType=13 AND ar.relatedProductId IS NOT NULL AND ar.relatedProductId!=0""")
    derived_by_volume = {}
    all_derived_pids = set()
    for src, der in cur.fetchall():
        derived_by_volume.setdefault(src, set()).add(der); all_derived_pids.add(der)

    der_type_by_pid = {}
    if all_derived_pids:
        cur.execute(f"SELECT productId, productType FROM Product WHERE productId IN ({','.join('?'*len(all_derived_pids))})", list(all_derived_pids))
        der_type_by_pid = {r[0]: r[1] for r in cur.fetchall()}

    all_pids = set()
    for grp in (horizons_products, markers_products, polygons_products, culture_products,
                volumes_products, workflow_products, velovly_products):
        all_pids.update(r[0] for r in grp)
    all_pids.update(all_derived_pids)
    attrs = attrs_for(list(all_pids), [AT_NAME, AT_DESC, AT_OWNER, AT_PATH, AT_CRS, AT_UNIT, AT_MEASURE])

    def ga(pid, at, default=""):
        v = attrs.get(pid, {}).get(at)
        return v if v is not None else default

    wb = Workbook()
    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    HDR_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    BORDER = Border(left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
                    top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"))
    ALIGN_TOP = Alignment(vertical="top", wrap_text=True)

    def write_sheet(title, headers, rows, col_widths=None, note=None):
        ws = wb.create_sheet(title); rn = 1
        if note:
            c = ws.cell(row=rn, column=1, value=note)
            c.font = Font(italic=True, color="7F6000"); c.fill = WARN_FILL
            ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=len(headers))
            rn += 1
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=rn, column=ci, value=h)
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[rn].height = 30; rn += 1
        for r in rows:
            for ci, v in enumerate(r, 1):
                c = ws.cell(row=rn, column=ci, value=v); c.alignment = ALIGN_TOP; c.border = BORDER
            rn += 1
        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A3" if note else "A2"
        return ws

    wb.remove(wb.active)

    cur.execute("SELECT key, value FROM ProjectMetadata")
    pm = dict(cur.fetchall())

    readme_rows = [
        ["Source", args.db], ["DB size", f"{os.path.getsize(args.db)/1e6:.1f} MB"],
        ["Project UUID", pm.get("projectUUID", "")],
        ["Schema", f"{pm.get('schemaVersion', '?')}.{pm.get('schemaMinorVersion', '?')}"],
        ["Generated", dt.datetime.utcnow().strftime("%Y-%b-%d %H:%M UTC")],
        ["", ""],
        ["METHOD", "Read project.dugprj as SQLite read-only. Aggregated Product+AttributeRevision rows. productType->category mapping INFERRED from filenames/sample data, not from documented DUG schema."],
        ["CONFIDENCE", "Counts are exact. Category labels are best-effort inference. See ProductType_Map sheet for the full mapping with caveats."],
        ["PROCESSES PER VOLUME — definition", "Step 1 (this sheet) shows the lineage links only — derived child products linked via AttributeRevision.relatedProductId. Step 2 (02_extract_parameters.py) enriches with full param spec. Step 3 (03_resolve_uuids.py) resolves UUIDs to product names."],
        ["", ""], ["COUNTS", ""],
        ["  Horizons (type 12)", len(horizons_products)],
        ["  Markers / pick names (type 8)", len(markers_products)],
        ["  Polygons (type 18)", len(polygons_products)],
        ["  Culture shapefiles (type 37)", len(culture_products)],
        ["  Volumes (type 13)", len(volumes_products)],
        ["  Workflow products (type 17)", len(workflow_products)],
        ["  Velocity/inversion overlays (type 52)", len(velovly_products)],
        ["  Distinct volume_headers UUIDs", len(vh_by_uuid)],
        ["  Total volume_headers rows (incl 2D lines)", sum(r[2] for r in vh_by_uuid.values())],
    ]
    ws = wb.create_sheet("README"); rn = 1
    for k, v in readme_rows:
        ws.cell(row=rn, column=1, value=k).font = Font(bold=bool(k) and not k.startswith("  "))
        ws.cell(row=rn, column=2, value=v).alignment = ALIGN_TOP
        rn += 1
    ws.column_dimensions["A"].width = 38; ws.column_dimensions["B"].width = 110

    horizons_rows = []
    for r in horizons_products:
        pid = r["productId"]
        horizons_rows.append([pid, "Horizon (gridded)", ga(pid, AT_NAME), r["owner"] or "",
                              ga(pid, AT_PATH), ga(pid, AT_CRS), ga(pid, AT_DESC)])
    for r in markers_products:
        pid = r["productId"]
        horizons_rows.append([pid, "Marker / pick name", ga(pid, AT_NAME), r["owner"] or "", "", "", ""])
    write_sheet("Horizons",
                ["productId","kind","name","owner","file path (if any)","CRS","description"],
                horizons_rows, [12,22,55,18,60,15,55],
                f"{len(horizons_products)} gridded horizons (type 12) + {len(markers_products)} marker/pick names (type 8).")

    polygon_rows = []
    for r in polygons_products:
        pid = r["productId"]
        polygon_rows.append([pid, "Polygon / contour / outline", ga(pid, AT_NAME),
                             r["owner"] or "", ga(pid, AT_CRS), ga(pid, AT_DESC)])
    for r in culture_products:
        pid = r["productId"]
        polygon_rows.append([pid, "Culture shapefile", ga(pid, AT_NAME),
                             r["owner"] or "", ga(pid, AT_CRS), ga(pid, AT_PATH)])
    write_sheet("Polygons",
                ["productId","kind","name","owner","CRS","description / source"],
                polygon_rows, [12,28,55,18,15,80],
                f"{len(polygons_products)} polygons (type 18) + {len(culture_products)} imported shapefiles (type 37).")

    volume_rows = []
    for r in volumes_products:
        pid = r["productId"]; uuid = vol_uuid_by_pid.get(pid)
        vh = vh_by_uuid.get(uuid)
        n_lines = vh[2] if vh else 0
        survey_kind = "3D" if vh and vh[2]==1 else ("2D survey (multi-line)" if vh and vh[2]>1 else "(no data)")
        rev = vol_rev.get(pid, {}); derived = derived_by_volume.get(pid, set())
        workflow_count = sum(1 for d in derived if der_type_by_pid.get(d)==17)
        velovly_count  = sum(1 for d in derived if der_type_by_pid.get(d)==52)
        volume_rows.append([pid, ga(pid, AT_NAME), r["owner"] or "", ga(pid, AT_PATH),
                            survey_kind, n_lines, vh[3] if vh else None, vh[4] if vh else None,
                            rev.get("n_rev",0), epoch_ms(rev.get("first_t")), epoch_ms(rev.get("last_t")),
                            rev.get("n_users",0), workflow_count, velovly_count, ga(pid, AT_DESC)])
    write_sheet("Volumes",
                ["productId","name","owner","file path","kind","# lines (vh)","trace_count","# samples (ns)",
                 "# revisions","first revision","last revision","# distinct users",
                 "# derived workflows (type17)","# vel/inv overlays (type52)","description"],
                volume_rows, [12,55,16,55,22,11,13,13,12,20,20,12,15,15,55],
                f"{len(volumes_products)} volume products (type 13). {len(vh_by_uuid)} have on-disk volume_headers; the rest are virtual/derived.")

    vp_rows = []
    for r in volumes_products:
        vpid = r["productId"]; vname = ga(vpid, AT_NAME)
        derived = sorted(derived_by_volume.get(vpid, set()))
        if not derived:
            vp_rows.append([vpid, vname, "—", "—", "(no derived processes link back to this volume)"])
            continue
        for dpid in derived:
            dtype = der_type_by_pid.get(dpid); cat = PT_MAP.get(dtype, ("?","?"))[0]
            dname = ga(dpid, AT_NAME) or "(unnamed)"; ddesc = ga(dpid, AT_DESC)
            vp_rows.append([vpid, vname, dpid, f"type {dtype}: {cat}",
                            dname + ((" — " + ddesc) if ddesc else "")])
    vp_rows.sort(key=lambda x: (str(x[1]), str(x[3]), str(x[4])))
    write_sheet("Volume_Processes",
                ["volume pid","volume name","derived pid","process type","process name / description"],
                vp_rows, [12,55,12,32,90],
                f"{sum(len(v) for v in derived_by_volume.values())} derived-product links across {len(derived_by_volume)} volumes. Step 2 will add a ProcessType column and full parameter string.")

    pt_rows = []
    cur.execute("SELECT productType, COUNT(*) FROM Product GROUP BY productType ORDER BY productType")
    for pt, n in cur.fetchall():
        cat, lbl = PT_MAP.get(pt, ("(unmapped)", "(no inference made)"))
        pt_rows.append([pt, n, cat, lbl])
    write_sheet("ProductType_Map",
                ["productType","# products","inferred category","notes / label"],
                pt_rows, [14,14,32,80],
                "INFERRED mapping. Not from documented DUG schema.")

    order = ["README","Volumes","Volume_Processes","Horizons","Polygons","ProductType_Map"]
    wb._sheets = [wb[s] for s in order]
    wb.save(args.out)
    print(f"OK step-1 -> {args.out}")
    print(f"  Horizons:        {len(horizons_rows)}")
    print(f"  Polygons:        {len(polygon_rows)}")
    print(f"  Volumes:         {len(volume_rows)}")
    print(f"  Volume_Processes:{len(vp_rows)}")


if __name__ == "__main__":
    main()
