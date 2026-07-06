"""
DUG project DB -> inventory xlsx (step 2 of 3).

Enriches Volume_Processes with the full per-process parameter spec from
AttributeRevision (attrType=528 for type-17 workflows; 25800/25802/25804/25805
for type-52 display overlays). Adds a deduplicated Process_Catalog sheet.

Edits the xlsx in place — no v2 files.

Usage:
    python 02_extract_parameters.py --db path/to/project.dugprj --out path/to/out.xlsx
"""
import argparse, sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    con = sqlite3.connect(f"file:{args.db}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row; cur = con.cursor()

    cur.execute("""SELECT DISTINCT ar.relatedProductId, ar.productId
                   FROM AttributeRevision ar JOIN Product src_p ON src_p.productId=ar.relatedProductId
                   WHERE src_p.productType=13 AND ar.relatedProductId IS NOT NULL AND ar.relatedProductId!=0""")
    edges = cur.fetchall()
    derived_pids = sorted({d for _, d in edges})
    if not derived_pids:
        print("no derived processes found — nothing to do")
        return

    ph = ",".join("?"*len(derived_pids))

    def bulk_attr(at):
        cur.execute(f"""SELECT productId, value FROM AttributeRevision
                        WHERE productId IN ({ph}) AND attributeType=? AND value IS NOT NULL""",
                    derived_pids+[at])
        out = {}
        for pid, v in cur.fetchall():
            if pid not in out or (not out[pid] and v):
                out[pid] = v
        return out

    names      = bulk_attr(1)
    desc       = bulk_attr(3)
    owner      = bulk_attr(4)
    proc_type  = bulk_attr(500)
    spec_528   = bulk_attr(528)
    ovly_kind  = bulk_attr(25800)
    ovly_tag   = bulk_attr(25802)
    ovly_cmap  = bulk_attr(25804)
    ovly_src   = bulk_attr(25805)

    cur.execute(f"SELECT productId, productType FROM Product WHERE productId IN ({ph})", derived_pids)
    ptype = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("""SELECT productId,
                          (SELECT value FROM AttributeRevision WHERE productId=p.productId AND attributeType=1 LIMIT 1) AS name
                   FROM Product p WHERE productType=13""")
    vol_name = {r[0]: r[1] for r in cur.fetchall()}

    def parse_528(text):
        """Parse the indented k/v spec -> ordered (key, value) pairs.
        First non-indented line is the workflow UUID; subsequent indented
        lines are key value or nested-block headers. Nested keys flatten
        with dot notation."""
        if not text: return [], None
        lines = text.splitlines()
        if not lines: return [], None
        workflow_uuid = lines[0].strip() if not lines[0].startswith(" ") else None
        pairs, stack = [], []
        for raw in (lines[1:] if workflow_uuid else lines):
            if not raw.strip(): continue
            indent = len(raw) - len(raw.lstrip(" "))
            body = raw.strip()
            while stack and indent <= stack[-1][0]:
                stack.pop()
            parts = body.split(None, 1)
            if len(parts) == 2:
                k, v = parts
                key_path = ".".join([s[1] for s in stack] + [k])
                pairs.append((key_path, v))
            else:
                stack.append((indent, body))
        return pairs, workflow_uuid

    def fmt_pairs(pairs, max_chars=900):
        s = "; ".join(f"{k}={v}" for k, v in pairs)
        return s if len(s) <= max_chars else s[:max_chars-3] + "..."

    edges_by_vol = {}
    for src, der in edges:
        edges_by_vol.setdefault(src, set()).add(der)

    cur.execute("SELECT productId FROM Product WHERE productType=13")
    all_vols = [r[0] for r in cur.fetchall()]

    rows_proc, proc_catalog, no_link = [], {}, 0
    for vpid in sorted(all_vols, key=lambda x: str(vol_name.get(x) or "")):
        vname = vol_name.get(vpid) or ""
        deriv = sorted(edges_by_vol.get(vpid, []))
        if not deriv:
            rows_proc.append([vpid, vname, "—", "—", "—", "(no derived processes link back to this volume)", ""])
            no_link += 1; continue
        for dpid in deriv:
            dt_ = ptype.get(dpid, "?")
            dname = names.get(dpid, "") or ""
            ddesc = desc.get(dpid, "") or ""
            ptlabel = proc_type.get(dpid, "")
            param_str = ""
            if dt_ == 17 and spec_528.get(dpid):
                pairs, _ = parse_528(spec_528[dpid])
                param_str = fmt_pairs(pairs)
            elif dt_ == 52:
                ko = []
                if ovly_kind.get(dpid):  ko.append(("kind", ovly_kind[dpid]))
                if ovly_cmap.get(dpid):  ko.append(("colormap", ovly_cmap[dpid]))
                if ovly_src.get(dpid):   ko.append(("source", ovly_src[dpid]))
                if ovly_tag.get(dpid):   ko.append(("tag", ovly_tag[dpid]))
                param_str = fmt_pairs(ko)
                ptlabel = ptlabel or "DisplayOverlay"
            elif spec_528.get(dpid):
                pairs, _ = parse_528(spec_528[dpid])
                param_str = fmt_pairs(pairs)

            rows_proc.append([vpid, vname, dpid, f"type {dt_}",
                              ptlabel or "(unspecified)",
                              (dname + ((" — " + ddesc) if ddesc else "")) or "(unnamed)",
                              param_str])
            if dpid not in proc_catalog:
                proc_catalog[dpid] = [dpid, f"type {dt_}", ptlabel or "(unspecified)",
                                       dname, ddesc, owner.get(dpid, ""), param_str]

    wb = load_workbook(args.out)
    for s in ("Volume_Processes", "Process_Catalog"):
        if s in wb.sheetnames: del wb[s]

    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    HDR_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    BORDER = Border(left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
                    top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"))
    ALIGN_TOP = Alignment(vertical="top", wrap_text=True)

    def write_at(name, idx, headers, rows, col_widths, note):
        ws = wb.create_sheet(name, idx)
        c = ws.cell(row=1, column=1, value=note)
        c.font = Font(italic=True, color="7F6000"); c.fill = WARN_FILL
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        for ci, h in enumerate(headers, 1):
            cc = ws.cell(row=2, column=ci, value=h)
            cc.font = HDR_FONT; cc.fill = HDR_FILL
            cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cc.border = BORDER
        ws.row_dimensions[2].height = 32
        for ri, r in enumerate(rows, start=3):
            for ci, v in enumerate(r, 1):
                cc = ws.cell(row=ri, column=ci, value=v); cc.alignment = ALIGN_TOP; cc.border = BORDER
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A3"
        return ws

    write_at("Volume_Processes", 2,
             ["volume pid","volume name","derived pid","derived type","ProcessType",
              "derived name / description","parameters (key=value; ...)"],
             rows_proc, [12,50,12,12,24,60,110],
             f"{len(rows_proc)} rows — one per (volume, derived process). Parameters from AttributeRevision attrType=528 (type-17) + 25800/25802/25804/25805 (type-52). UUIDs in parameters are raw (step 3 resolves them).")

    catalog_rows = sorted(proc_catalog.values(), key=lambda r: (str(r[2]), str(r[3])))
    write_at("Process_Catalog", 3,
             ["derived pid","derived type","ProcessType","name","description","owner","parameters"],
             catalog_rows, [12,12,24,50,50,16,110],
             f"{len(catalog_rows)} distinct derived process products. Browse the unique processes without volume cross-join.")

    ws_r = wb["README"]
    row = ws_r.max_row + 2
    ws_r.cell(row=row, column=1, value="UPDATE — parameter extraction").font = Font(bold=True, color="1F4E78"); row += 1
    ws_r.cell(row=row, column=1, value="  Volume_Processes rows (with parameters)").alignment = ALIGN_TOP
    ws_r.cell(row=row, column=2, value=len(rows_proc)).alignment = ALIGN_TOP; row += 1
    ws_r.cell(row=row, column=1, value="  Distinct derived processes (Process_Catalog)").alignment = ALIGN_TOP
    ws_r.cell(row=row, column=2, value=len(catalog_rows)).alignment = ALIGN_TOP; row += 1
    ws_r.cell(row=row, column=1, value="  Volumes with no derived process link").alignment = ALIGN_TOP
    ws_r.cell(row=row, column=2, value=no_link).alignment = ALIGN_TOP; row += 1
    distinct_pts = len({r[2] for r in catalog_rows if r[2] and r[2] != '(unspecified)'})
    ws_r.cell(row=row, column=1, value="  Distinct ProcessType labels").alignment = ALIGN_TOP
    ws_r.cell(row=row, column=2, value=distinct_pts).alignment = ALIGN_TOP

    desired = ["README","Volumes","Volume_Processes","Process_Catalog","Horizons","Polygons","ProductType_Map"]
    wb._sheets = [wb[n] for n in desired if n in wb.sheetnames]
    wb.save(args.out)
    print(f"OK step-2 -> {args.out}")
    print(f"  Volume_Processes: {len(rows_proc)}")
    print(f"  Process_Catalog:  {len(catalog_rows)}")
    print(f"  ProcessType labels: {distinct_pts}")
    print(f"  Volumes without link: {no_link}")


if __name__ == "__main__":
    main()
